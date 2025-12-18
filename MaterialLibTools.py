import autometrixZipTools as AZT
from icecream import ic
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import quote_sheetname, absolute_coordinate, get_column_letter
import RawMaterialFromOdoo

workbook_path=r"C:\AutometrixTools\Spexlib.xlsx"
worksheet="Spexlib"
ID="ID"
CNC_NAME="CNC Name"
SPACING="Gap between pieces"
WIDTH="Width"
ODOO_CODE="Odoo Code"
ODOO_NAME="Odoo Name"
ODOO_VENDOR_CODE="Vendor Code"
ODOO_UOM="UOM"
ODOO_VENDOR_NAME="Vendor Name"

ODOO_DATA_COLS=(
    (ODOO_CODE,5),
    (ODOO_NAME,6),
    (ODOO_VENDOR_CODE,7),
    (ODOO_UOM,8),
    (ODOO_VENDOR_NAME,9)
)

COL_HEADERS=[ID,CNC_NAME,SPACING,WIDTH]
COL_HEADERS_ODOO=[ODOO_NAME,ODOO_VENDOR_CODE,ODOO_UOM,ODOO_VENDOR_NAME]
RELEVANT_TAGS={"a2:MaterialID":ID,
               "a2:Name":CNC_NAME,
               "a2:CopySpacing":SPACING,
               "a2:MaterialPattern":WIDTH}

def get_materials_from_AMLib(lib_path= r"C:\Users\SimonCaldwell\Medifab\Spex CNC - MarkerFiles-PatternSmith\2025-08-27_Material.library")->list[dict]:
    #extract_to = r"C:\Users\SimonCaldwell\Desktop\extracted files\lib"
    groupsXml=AZT.unzip_specific_XML_file(lib_path,"Groups.xml",)
    if groupsXml:
        MatData=AZT.get_specific_XML_data(groupsXml,"a2:MaterialTemplate",AZT.RELEVANT_TAGS.keys())
    else:
        print(groupsXml)
    return MatData
material_data=get_materials_from_AMLib()
# switch from Autometrix tag names to names defined in RELEVANT TAGS
material_data=[{RELEVANT_TAGS[k]:v for k, v in material.items()} for material in material_data]
ic(material_data)

def update_Odoo_data(ws:Worksheet):
    rm=ws.iter_rows()
    data_refs=[]
    first_row=next(rm)
    #([v.value for v in first_row])
    for row in rm:
        if row[0].value==None:
            continue
        existing_data_ref=({first_row[i].value:col for  i,col in enumerate(row)})
        data_refs.append(existing_data_ref)
        if existing_data_ref[ODOO_CODE].value==None:
            continue
        new_odoo_data=RawMaterialFromOdoo.odoo_import([existing_data_ref[ODOO_CODE].value])
        #new_odoo_data=[v for k,v in new_odoo_data[0].items()]
        ic(new_odoo_data)
        current_row=row[0].row
        if len(new_odoo_data)<1:
            continue
        for i,data_value in enumerate(new_odoo_data[0]):
            ws.cell(row=current_row,column=i+6).value=data_value
        
        #old_odoo_data=rm.

wb = load_workbook(filename = workbook_path)
ws=wb.active
existing_material_refs=[]
row_machine=ws.iter_rows()
first_row=next(row_machine)
#([v.value for v in first_row])
for row in row_machine:
    if row[0].value==None:
        continue
    ic(row[0].value)
    existing_material_entry={#convert numerical values to 1 decimal place because this dict is used to compare with autometrix values
                            first_row[i].value:"{:.1f}".format(float(col.value))if str(col.value).isdigit()  else str(col.value)
                            for  i,col in enumerate(row)
                            #exclude all the cols that dont match info coming from autometix 
                            if first_row[i].value in COL_HEADERS
                            }
    #make a generator of dict Header:<cell Reference>
    existing_material_ref=({first_row[i].value:col for  i,col in enumerate(row)})
    #add generator to list
    existing_material_refs.append(existing_material_ref)
    # find relvant existing data 
    try:
        new_material_match=next(d for d in material_data if d[ID]==existing_material_entry[ID])
    except:new_material_match=None #material is not in existing sheet. 

    if not existing_material_entry==new_material_match and not new_material_match==None:
        ic(existing_material_entry,new_material_match)
        workbook_changed=True
        for k,v in new_material_match.items():
            existing_material_ref[k].value=v
    elif new_material_match == None:
        ic(existing_material_entry,"is no longer in the Library")

    existing_mat_ids=[str(mat[ID].value) for mat in existing_material_refs]
    
for auto_lib_mat in material_data:
    if not auto_lib_mat[ID] in existing_mat_ids:
        ic("new mat",auto_lib_mat)
        ws.append((v for k,v in auto_lib_mat.items()))

Odoo_Codes=(d[ODOO_CODE].value for d in existing_material_refs if not d[ODOO_CODE].value == None)
Odoo_data= RawMaterialFromOdoo.odoo_import(list(Odoo_Codes))
update_Odoo_data(ws)

wb.save(workbook_path)
#    ic(row)
#    pass
'''for i,var in enumerate(varData):
    #ic(f"A{i+1}")
    #ws[f"A{i+1}"]=var["Name"]
    ws.cell(1,i+1).value=var[NAME]
    ws.cell(2,i+1).value=f"{var[EXPR]}"
    ws.cell(3,i+1).value=f"={var[EXPR]}"
    # make sure sheetnames  and cell referencesare quoted correctly
    letter=get_column_letter(i+1)
    ref = f"{quote_sheetname(ws.title)}!{absolute_coordinate(f'{letter}')}$3"

    defn = DefinedName(var[NAME], attr_text=ref,publishToServer=True)
    ws.defined_names.add(defn)
'''
wb.close()
#AZT.extract_zip_file(file_path,extract_to)
