
import autometrixZipTools as AZT

from icecream import ic
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate, get_column_letter
import re
import svg
from random import choice
svgPath=r"C:\binary\tem.svg"
rows = 4
columns = 6
width = 60
height = 40
radius = 4
colors = [
    '#2ecc71',
    '#3498db',
    '#9b59b6',
    '#34495e',
    '#e67e22',
    '#e74c3c',
    '#7f8c8d',
]
XPOS="x"
YPOS="y"
INCHES="Inches"
CM="Centimeters"
MM="Millimeters"
CONVERT_TO_MM_FACTORS={INCHES:25.4,
                       CM:10,
                       MM:1}
# UNIT_CONVER[units from current param][units of referenced param] 
UNITS_CONVERT={MM:{MM:1,
                   CM:0.1,
                   INCHES:25.4},
                CM:{MM:10,
                    CM:1,
                    INCHES:2.54},
                INCHES:{MM:0.0394,
                        CM:0.394,
                        INCHES:1}
                        }
EXPR="Expression"
NAME="Name"
UNITS="Units"
MEAS="UsesMeasured"
HASVALUE="HasValue"
TYPE="Type"
TIER="tier"
CONTAINS="contains"

TYPE_MEASURED="Measured"

re_digits_in_exp=r"(\d+\.\d+|\d+)"
re_words_in_eq=r"([a-zA-Z]+)"
RelvantFeilds=[EXPR,
               NAME,
               UNITS,
               TYPE]


def _getTier(var:dict,varDict:dict,tier=0)->int:
    tier+=1
    ic((var[NAME]))
    if CONTAINS in var.keys() and len(var[CONTAINS])==0:
        return tier
    tierhold=tier
    
    for i,vie in enumerate(var[CONTAINS]):
        ic(vie,tier,i)
        ic(_getTier(VarDict[vie],varDict,tier=tier),tier)
        tier=(max(_getTier(VarDict[vie],varDict,tier=tier),tier))
        tierhold=max(tier,tierhold)
        tier=1
    return tierhold
    
def convert_Vars_to_units(parmeter:dict,otherParam:list[dict])->dict:
    #function is inside so it can use otherParam
    
    def _re_sub_multiply_by(x):
        
        x=x.group(0) 
        ic(len(otherParam))
        for op in otherParam:
            ic(x,op[NAME])
            if op[NAME]==x:
                ic(UNITS_CONVERT[parmeter[UNITS]][op[UNITS]])
                factor=UNITS_CONVERT[parmeter[UNITS]][op[UNITS]]
                if factor==1:
                    return x # no reaon to add *1
                return f'{x}*{factor}'
        else: return x # var not found, might be an automtrix builtin
    if parmeter[TYPE]==TYPE_MEASURED:
        parmeter[EXPR]="MEASUREDLINE"
        # for measured types Expression is a id for the line to be measured bad to convert that.
        return parmeter
    #paramUnitsFactor=CONVERT_TO_MM_FACTORS[parmeter[UNITS]]
    ic(parmeter[EXPR])
    parmeter[EXPR]=re.sub(re_words_in_eq,_re_sub_multiply_by,parmeter[EXPR])
    ic(parmeter[EXPR])
    return parmeter

XMLData=AZT.unzip_specific_XML_file(r"C:\patternsmith\1255 backrest.templates",r"Templates/Variables")
parmData=AZT.get_specific_XML_data(XMLData,"KeyValueOfstringVariableKLLHlUQr",RelvantFeilds,convertValues=False)
XMLData2=AZT.unzip_specific_XML_file(r"C:\patternsmith\1255 backrest.templates",r"Templates/BasicData")
varData=AZT.get_specific_XML_data(XMLData2,"KeyValueOfstringVariableKLLHlUQr",RelvantFeilds,convertValues=False)
ic(varData)
varData=varData+parmData
wb=Workbook()
ws=wb.active

for i,var in enumerate(varData):
    #ic(f"A{i+1}")
    #ws[f"A{i+1}"]=var["Name"]
    ws.cell(1,i+1).value=var[NAME]
    ws.cell(2,i+1).value=f"{var[EXPR]}"
    var=convert_Vars_to_units(var,varData)
    ws.cell(3,i+1).value=f"={var[EXPR]}"
    # make sure sheetnames  and cell referencesare quoted correctly
    letter=get_column_letter(i+1)
    ref = f"{quote_sheetname(ws.title)}!{absolute_coordinate(f'{letter}')}$3"

    defn = DefinedName(var[NAME], attr_text=ref,publishToServer=True)
    ws.defined_names.add(defn)

   
print(ws.defined_names)

wb.save("backrest_template_vars.xlsx")

VarDict={D[NAME]:D for D  in varData}
ic(VarDict)
for var in varData:
    varsInEXPR=re.findall(re_words_in_eq,var[EXPR])
    varsInEXPR=[v for v in varsInEXPR if v in VarDict.keys() and v is not var[NAME]]
    varsInEXPR=list(set(varsInEXPR))
    var[CONTAINS]=varsInEXPR
    var[TIER]=min(len(varsInEXPR)+1,3)
    ordered_Tem_Var={}

for tem_Var in varData:
    tem_Var[TIER]=_getTier(tem_Var,VarDict)
    if str(tem_Var[TIER]) in ordered_Tem_Var:
        ordered_Tem_Var[str(tem_Var[TIER])]+=[tem_Var]
    else:
        ordered_Tem_Var[str(tem_Var[TIER])]=[tem_Var]
ordered_Tem_Var=dict(sorted(ordered_Tem_Var.items()))
ic(ordered_Tem_Var)


def random_circles(x, y):
    yield svg.Circle(
        cx=x, cy=y, r=radius,
        fill=choice(colors),
    )
    yield svg.Circle(
        cx=x, cy=y, r=radius // 2,
        fill='none',
        stroke='white',
        stroke_width=1,
    )
def TemVarNameText(TemVar:dict):
    yield svg.Text(
        text=TemVar[NAME],
        font_size=1,
        x=TemVar[XPOS],
        y=TemVar[YPOS]
        )
    

    


circles = []
step_x = width // columns
step_y = height // rows
for y, vl in enumerate(ordered_Tem_Var.items()):
    ic(y,vl[1])
    for x, tem_var in enumerate(vl[1]):
        tem_var.update({XPOS:x*step_x+10, YPOS:y*step_y})
        circles.extend(random_circles(tem_var[XPOS],tem_var[YPOS]))
        circles.extend(TemVarNameText(tem_var))
        ic(x,y)
        ic(tem_var)

canvas = svg.SVG(
    viewBox=svg.ViewBoxSpec(0, 0, width, height),
    elements=circles,
)

with open(svgPath,"w+") as SVGFile:
    SVGFile.write(canvas.as_str())